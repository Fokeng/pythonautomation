import pandas as pd
import openpyxl  # library that loads and reads .xlsx files
import csv      # library that reads and writes .csv files
from openpyxl import load_workbook 
from openpyxl.utils.dataframe import dataframe_to_rows

wb = load_workbook('Employee_data.xlsx') # load the Employee_data.xlsx file

ws = wb.active                          # loading the active worksheet from Employee_data.xlsx 

df = pd.DataFrame(ws.values)          # storing the data from the worksheet in a dataframe

df1 = pd.read_csv("Employee_data.csv") # reading the .csv file and storing it in a dataframe 

'''
formating the data in column 2 at index 1 of the dataframe
'''
df[1].replace("helpinghands.cm", "handsinhands.org",regex=True,inplace=True)

'''
formating the data under the Email address column in the dataframe
'''
df1['Email address'].replace("helpinghands.cm", "handsinhands.org",regex=True,inplace=True)

# creating a new workbook to store the updated .xlsx file
wb2 = openpyxl.Workbook() 
ws = wb2.active
for r in dataframe_to_rows(df, index=True, header= False):
    ws.append(r)
wb2.save('Employee_data_update.xlsx') # saving the workbook as a .xlsx file

df1.to_csv('Employee_data_update.csv',index = False) # saving the updated data in the dataframe as a .csv file


